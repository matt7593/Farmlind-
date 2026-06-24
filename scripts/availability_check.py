import json
import base64
import urllib.request
import urllib.parse
import io
import os
from datetime import datetime
from collections import defaultdict

import anthropic
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

SLACK_BOT_TOKEN = os.environ["SLACK_BOT_TOKEN"]
ANTHROPIC_API_KEY = os.environ["ANTHROPIC_API_KEY"]
GMAIL_CLIENT_ID = os.environ["GMAIL_CLIENT_ID"]
GMAIL_CLIENT_SECRET = os.environ["GMAIL_CLIENT_SECRET"]
GMAIL_REFRESH_TOKEN = os.environ["GMAIL_REFRESH_TOKEN_1"]

ALL_EMAILS  = ["matt@farmlindproduce.com", "farmlindproduce@gmail.com", "babbe331@gmail.com", "howard@farmlindproduce.com"]
TEST_EMAILS = ["babbe331@gmail.com"]
NOTIFY_EMAILS = TEST_EMAILS if os.environ.get("TEST_MODE") == "true" else ALL_EMAILS
CHANNEL_NAME = "availability-list-changes"
REFERENCE_FILE = os.path.join(os.path.dirname(__file__), "product_reference.xlsx")

claude = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY, timeout=120.0, max_retries=2)


# ── Product reference loader ───────────────────────────────────────────────────

def load_product_reference():
    """
    Load known product names from the internal availability spreadsheet.
    Returns a deduplicated list of product name strings Claude can use
    to map vendor abbreviations to real names.
    """
    if not os.path.exists(REFERENCE_FILE):
        print("  No product_reference.xlsx found — skipping reference lookup")
        return []

    names = set()
    try:
        wb = openpyxl.load_workbook(REFERENCE_FILE, read_only=True, data_only=True)

        # List sheet col A — full product names used internally
        if "List" in wb.sheetnames:
            ws = wb["List"]
            for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
                val = row[0]
                if val and isinstance(val, str):
                    val = val.strip().replace("\xa0", "")
                    # Skip section headers (all caps) and empty
                    if val and not val.isupper():
                        names.add(val)

            # Reserva Variants col 35 (AJ) — official system names
            for row in ws.iter_rows(min_col=35, max_col=35, values_only=True):
                val = row[0]
                if val and isinstance(val, str):
                    val = val.strip()
                    if val and val not in ("Reserva names",):
                        names.add(val)

        # Product IDs sheet — display_name column
        if "Product IDs" in wb.sheetnames:
            ws2 = wb["Product IDs"]
            for row in ws2.iter_rows(min_col=2, max_col=2, values_only=True):
                val = row[0]
                if val and isinstance(val, str) and val.strip() != "display_name":
                    names.add(val.strip())

    except Exception as e:
        print(f"  Warning: could not load product reference: {e}")

    print(f"  Loaded {len(names)} product names from reference file")
    return sorted(names)


# ── Slack helpers ──────────────────────────────────────────────────────────────

def slack_get(path, params=None):
    url = f"https://slack.com/api{path}"
    if params:
        url += "?" + urllib.parse.urlencode(params)
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {SLACK_BOT_TOKEN}"})
    with urllib.request.urlopen(req) as resp:
        data = json.loads(resp.read())
    if not data.get("ok"):
        raise RuntimeError(f"Slack API error on {path}: {data.get('error')}")
    return data


def find_channel_id(name):
    cursor = None
    while True:
        params = {"limit": 200, "exclude_archived": "true", "types": "public_channel,private_channel"}
        if cursor:
            params["cursor"] = cursor
        data = slack_get("/conversations.list", params)
        for ch in data.get("channels", []):
            if ch["name"] == name:
                return ch["id"]
        cursor = data.get("response_metadata", {}).get("next_cursor")
        if not cursor:
            break
    raise RuntimeError(f"Channel #{name} not found — make sure the bot is invited to it")


def fetch_channel_history(channel_id, limit=200):
    data = slack_get("/conversations.history", {"channel": channel_id, "limit": limit})
    return data.get("messages", [])


def has_messages_today(messages):
    """Return True if any message was posted today (ET)."""
    from datetime import timezone, timedelta
    et = timezone(timedelta(hours=-4))  # EDT; use -5 for EST
    today = datetime.now(et).date()
    for msg in messages:
        ts = float(msg.get("ts", 0))
        msg_date = datetime.fromtimestamp(ts, tz=et).date()
        if msg_date == today:
            return True
    return False


def get_user_display_name(user_id):
    try:
        data = slack_get("/users.info", {"user": user_id})
        profile = data.get("user", {}).get("profile", {})
        return profile.get("display_name") or profile.get("real_name") or user_id
    except Exception:
        return user_id


def download_slack_file(url):
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {SLACK_BOT_TOKEN}"})
    with urllib.request.urlopen(req) as resp:
        return resp.read()


# ── Claude helpers ─────────────────────────────────────────────────────────────

def build_extract_prompt(product_names):
    reference_section = ""
    if product_names:
        sample = product_names[:300]
        reference_section = (
            "\n\nKNOWN PRODUCT NAMES (use these to resolve abbreviations and shorthand):\n"
            + "\n".join(f"- {n}" for n in sample)
            + "\n\nWhen you see an abbreviation or short name, match it to the closest known product name above."
        )

    return f"""You are parsing a produce availability/price list. Extract the vendor name(s) and all priced items.

VENDOR NAME RULES — identify the vendor using these exact mappings:
- "TMK", "Tmk" → "TMK"
- "Aurpack", "Aurback", "Auerbach", "Aurebach", "Auerpak", "Aureback" → "Aurpack"
- "Nathel", "Nathel & Nathel", "Nathel and Nathel", "Nathel list", "Nathel prices", "PD371" → "Nathel"
- "Top line", "Top Line" → "Top Line"
- "A and j", "A and J", "A & J", "A&J" → "A & J Produce Corp"
- "Dagele", "DAGELE" → "Dagele Brothers"
- "Triple J" → "Triple J"
- "Andy boy", "Andy Boy" → "Andy Boy"
- "Stews", "Stew", "Stew Leonard" → "Stew Leonards"
- "Dottavio", "M Dottavio", "M. Dottavio", "Dottavio Produce", "mdottavioproduce" → "Dottavio"
- "Armotta", "Aramotta", "Armotta Produce" → "Armotta"
- "Donaldson", "Donaldson Farms" → "Donaldson"
- "Cassaday", "Cassady", "Cassaday Farms" → "Cassaday"
- "Tranquility", "Tranquillity" → "Tranquility"
- "Landisville", "Landisville Produce" → "Landisville"
- A Nathel order/screenshot labeled with "Josh" (Josh's order/list) → "Josh Nathel"
- A Nathel order/screenshot labeled with "Paul" (Paul's order/list) → "Paul Nathel"
- SPECIALS lists (e.g. "Sunday Specials", "Monday Specials", ... any day Sunday-Saturday,
  or "<day> prices/specials"): set the vendor to "<Day> Specials - <Supplier>". You MUST
  attach the supplier/company offering the specials (the vendor name in the message),
  normalized using the mappings above — e.g. "Tuesday Specials - TMK",
  "Saturday Specials - Aurpack".
  SUNDAY specials are ALWAYS offered by either TMK or Nathel — determine which one from
  the content and label it "Sunday Specials - TMK" or "Sunday Specials - Nathel".
  Only omit the supplier (use plain "<Day> Specials") if truly no supplier can be determined.
- NEVER use "Farmlind", "Farmlind Produce", "Matt", "Matt Lind", or "Hunts Point" as a vendor — they are
  the BUYER purchasing from these vendors (or a market reference), not a seller. If the only name present is
  one of these, return null for vendor.
- If no vendor can be identified → return null for vendor

MULTIPLE VENDORS IN ONE MESSAGE: If the content contains separate lists for more than
one distinct vendor or person (for example a Nathel screenshot showing both "Josh" and
"Paul" orders, or two suppliers' lists stacked together), return each as its own entry in
the "groups" array, each with its own vendor name and items. Otherwise leave "groups" empty
and just fill "vendor" and "items".

Return ONLY a JSON object:
{{
  "vendor": "<vendor name or null>",
  "items": [
    {{
      "item": "<produce item name ONLY — no sizes, counts, or grades — in Title Case>",
      "price": <price as a number only, no $ sign>,
      "unit": "<size/count/weight/grade — e.g. '72ct', '36ct', '2.5 inch', '5 pack', '50lb', '12/3lb bags', 'case', 'lb', 'bunch' — include ALL distinguishing size or count info here>"
    }}
  ],
  "groups": [
    {{ "vendor": "<vendor name>", "items": [ {{ "item": ..., "price": ..., "unit": ... }} ] }}
  ]
}}

Rules:
- Item name must be the CLEAN product name only — NO counts, sizes, or grades in the name. Put all of that in the unit field instead.
  - WRONG: "Fuji Apple 72ct", "Grapefruit 36ct", "Lime 230ct"
  - RIGHT: item="Fuji Apple", unit="72ct" / item="Grapefruit", unit="36ct" / item="Lime", unit="230ct"
- ALWAYS write the variety/descriptor BEFORE the base fruit/vegetable name (e.g., "Fuji Apple" NOT "Apple Fuji", "Navel Orange" NOT "Orange Navel", "Hass Avocado" NOT "Avocado Hass", "Roma Tomato" NOT "Tomato Roma", "Bartlett Pear" NOT "Pear Bartlett").
- Use SINGULAR form for all item names (e.g., "Strawberry" not "Strawberries", "Avocado" not "Avocados", "Lime" not "Limes").
- If the same item appears at multiple prices, put the distinguishing size/count/grade in the unit field so they can be told apart.
- ALWAYS populate the "unit" field for EVERY item. Capture any count, size, pack, or weight
  shown — e.g. "72ct", "36ct", "12/3lb", "50lb", "5 pack", "18 oz", "1 lb", "bunch", "12ct".
  If the list shows no explicit size/count for an item, infer the natural selling unit
  ("case", "lb", "bunch", "each") rather than leaving it blank. Never return an empty unit.
- Price must be a number only — no $ symbol, no slashes.
- If a price range is given (e.g. 1.00-1.50), use the lower number.
- Do NOT include items with no price.
- Return valid JSON only — no markdown, no explanation.{reference_section}"""


def extract_prices_from_content(content_blocks, sender_name, product_names):
    prompt = build_extract_prompt(product_names)
    result = claude.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=8192,
        messages=[{
            "role": "user",
            "content": content_blocks + [{"type": "text", "text": prompt}]
        }]
    )
    raw = result.content[0].text.strip()
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        start, end = raw.find("{"), raw.rfind("}")
        if start != -1 and end != -1:
            try:
                data = json.loads(raw[start:end + 1])
            except Exception:
                print(f"  Could not parse Claude response: {raw[:200]}")
                return None
        else:
            return None

    # Skip messages where Claude couldn't identify a real vendor name
    if not data.get("vendor") or data["vendor"] in ("Unknown Vendor", ""):
        data["vendor"] = None

    return data


def pdf_page_texts(file_bytes):
    """Return a list of text strings, one per PDF page."""
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(file_bytes))
    return [(page.extract_text() or "") for page in reader.pages]


# Cheap, no-API vendor detection from raw text — used to skip older duplicate PDFs
# without paying for a Claude call. Keyword (lowercase) -> canonical vendor name.
VENDOR_KEYWORDS = [
    ("nathel & nathel", "Nathel"),
    ("nathel", "Nathel"),
    ("pd371", "Nathel"),
    ("aurpack", "Aurpack"),
    ("auerpak", "Aurpack"),
    ("auerbach", "Aurpack"),
    ("aurback", "Aurpack"),
    ("a & j produce", "A & J Produce Corp"),
    ("dagele", "Dagele Brothers"),
    ("triple j", "Triple J"),
    ("andy boy", "Andy Boy"),
    ("stew leonard", "Stew Leonards"),
    ("top line", "Top Line"),
    ("dottavio", "Dottavio"),
    ("armotta", "Armotta"),
    ("aramotta", "Armotta"),
    ("donaldson", "Donaldson"),
    ("cassaday", "Cassaday"),
    ("tranquility", "Tranquility"),
    ("landisville", "Landisville"),
]


def guess_vendor_from_text(text):
    """Best-effort vendor name from raw text using known keywords. Returns None if
    no confident match (caller should then fall back to the model)."""
    low = text.lower()
    for kw, canonical in VENDOR_KEYWORDS:
        if kw in low:
            return canonical
    return None


def extract_prices_from_pdf_chunked(file_bytes, product_names, pages_per_chunk=2,
                                    seen_check=None):
    """Extract prices from a large multi-page PDF by processing it in page-chunks,
    so a long list (hundreds of items) is never truncated by the model's output limit.

    seen_check: optional callable(vendor) -> bool. Called as soon as the vendor is
    identified (after the first chunk). If it returns True, the vendor has already been
    processed from a fresher source, so we stop immediately and return None — avoiding
    dozens of wasted API calls on duplicate/older lists.

    Returns a single merged {"vendor": ..., "items": [...]} dict, or None if skipped."""
    pages = pdf_page_texts(file_bytes)
    if not pages or not any(p.strip() for p in pages):
        # No extractable text (scanned PDF) — fall back to sending the whole document
        block = {
            "type": "document",
            "source": {"type": "base64", "media_type": "application/pdf",
                       "data": base64.b64encode(file_bytes).decode()}
        }
        return extract_prices_from_content([block], "PDF", product_names)

    # No-API shortcut: if we can recognize the vendor from the raw text and that
    # vendor was already handled, skip the whole PDF without any Claude calls.
    if seen_check:
        guessed = guess_vendor_from_text("\n".join(pages[:2]))
        if guessed and seen_check(guessed):
            print(f"    Vendor {guessed} already processed (text match) — skipping PDF, no API calls")
            return None

    vendor = None
    all_items = []
    total_chunks = (len(pages) + pages_per_chunk - 1) // pages_per_chunk
    for i in range(0, len(pages), pages_per_chunk):
        chunk_text = "\n".join(pages[i:i + pages_per_chunk]).strip()
        if not chunk_text:
            continue
        chunk_num = i // pages_per_chunk + 1
        print(f"    Chunk {chunk_num}/{total_chunks} (pages {i + 1}-{min(i + pages_per_chunk, len(pages))})...")
        try:
            result = extract_prices_from_content(
                [{"type": "text", "text": chunk_text}], "PDF", product_names
            )
        except Exception as e:
            print(f"      Chunk error: {e}")
            continue
        if not result:
            continue
        if not vendor and result.get("vendor"):
            vendor = result["vendor"]
            # Bail out early if this vendor was already handled (e.g. an older
            # duplicate post). Saves chunk-reading the rest of the PDF.
            if seen_check and seen_check(vendor):
                print(f"    Vendor {vendor} already processed — skipping rest of PDF")
                return None
        all_items.extend(result.get("items", []))

    return {"vendor": vendor, "items": all_items}


# ── Gmail helpers ──────────────────────────────────────────────────────────────

GMAIL_REFRESH_TOKEN_2 = os.environ.get("GMAIL_REFRESH_TOKEN_2")  # farmlindproduce@gmail.com


def get_access_token(refresh_token=None):
    data = urllib.parse.urlencode({
        "client_id": GMAIL_CLIENT_ID,
        "client_secret": GMAIL_CLIENT_SECRET,
        "refresh_token": refresh_token or GMAIL_REFRESH_TOKEN,
        "grant_type": "refresh_token"
    }).encode()
    req = urllib.request.Request("https://oauth2.googleapis.com/token", data=data)
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read())["access_token"]


def gmail_get(access_token, path, params=None):
    url = f"https://gmail.googleapis.com/gmail/v1{path}"
    if params:
        url += "?" + urllib.parse.urlencode(params)
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {access_token}"})
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read())


def _gmail_extract_text(part):
    """Recursively pull the text/plain body from a Gmail message payload."""
    mime = part.get("mimeType", "")
    body_data = part.get("body", {}).get("data", "")
    if mime == "text/plain" and body_data:
        return base64.urlsafe_b64decode(body_data + "==").decode("utf-8", errors="ignore")
    for sub in part.get("parts", []):
        txt = _gmail_extract_text(sub)
        if txt:
            return txt
    # Fall back to HTML if no plain part
    if mime == "text/html" and body_data:
        import re as _re
        html = base64.urlsafe_b64decode(body_data + "==").decode("utf-8", errors="ignore")
        return _re.sub(r"<[^>]+>", " ", html)
    return ""


def fetch_latest_dottavio_email(days=10):
    """Search the Gmail accounts for the most recent Dottavio price-list email.
    Returns (body_text, internal_date_ms) or (None, None) if not found."""
    accounts = []
    if GMAIL_REFRESH_TOKEN_2:
        accounts.append(("farmlindproduce@gmail.com", GMAIL_REFRESH_TOKEN_2))
    accounts.append(("matt@farmlindproduce.com", GMAIL_REFRESH_TOKEN))

    query = f"from:(dottavio OR mdottavioproduce.com) newer_than:{days}d"
    best_text, best_ts = None, -1
    for email_addr, refresh in accounts:
        try:
            token = get_access_token(refresh)
            res = gmail_get(token, "/users/me/messages", {"q": query, "maxResults": 10})
            for m in res.get("messages", []):
                full = gmail_get(token, f"/users/me/messages/{m['id']}", {"format": "full"})
                ts = int(full.get("internalDate", 0))
                if ts > best_ts:
                    text = _gmail_extract_text(full.get("payload", {}))
                    if text.strip():
                        best_text, best_ts = text, ts
        except Exception as e:
            print(f"  Dottavio email fetch error ({email_addr}): {e}")
    return best_text, (best_ts if best_ts >= 0 else None)


def availability_already_sent_today(today_et):
    """True if an Availability Price Comparison email was already sent today (ET),
    checked in matt@'s Sent mail. Prevents the Thursday watch from sending repeatedly."""
    try:
        token = get_access_token()  # matt@ (sender)
        after = today_et.strftime("%Y/%m/%d")
        res = gmail_get(token, "/users/me/messages", {
            "q": f'in:sent subject:"Availability Price Comparison" after:{after}',
            "maxResults": 5
        })
        return len(res.get("messages", [])) > 0
    except Exception as e:
        print(f"  Sent-check error: {e}")
        return False


def send_email_with_attachment(access_token, subject, body_text, xlsx_bytes):
    attachment_b64 = base64.b64encode(xlsx_bytes).decode()
    filename = f"availability_{datetime.now().strftime('%Y%m%d')}.xlsx"
    boundary = "==farmlind_boundary=="
    recipients = ", ".join(NOTIFY_EMAILS)

    mime = (
        f"From: matt@farmlindproduce.com\r\n"
        f"To: {recipients}\r\n"
        f"Subject: {subject}\r\n"
        f"MIME-Version: 1.0\r\n"
        f"Content-Type: multipart/mixed; boundary=\"{boundary}\"\r\n\r\n"
        f"--{boundary}\r\n"
        f"Content-Type: text/plain; charset=utf-8\r\n\r\n"
        f"{body_text}\r\n\r\n"
        f"--{boundary}\r\n"
        f"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n"
        f"Content-Transfer-Encoding: base64\r\n"
        f"Content-Disposition: attachment; filename=\"{filename}\"\r\n\r\n"
        f"{attachment_b64}\r\n\r\n"
        f"--{boundary}--"
    )
    encoded = base64.urlsafe_b64encode(mime.encode()).decode()
    url = "https://gmail.googleapis.com/gmail/v1/users/me/messages/send"
    req_data = json.dumps({"raw": encoded}).encode()
    req = urllib.request.Request(
        url, data=req_data,
        headers={"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"}
    )
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read())


# ── Email body builder ─────────────────────────────────────────────────────────

def build_email_body(item_vendor_map, today_str):
    lines = [
        f"Availability Price Comparison — {today_str}",
        "=" * 50,
        "",
        "Items are sorted alphabetically.",
        "Vendors listed cheapest to most expensive.",
        "",
        "=" * 50,
        "",
    ]

    for item in sorted(item_vendor_map.keys()):
        vendors = item_vendor_map[item]  # list of (vendor, price, unit)
        vendors_sorted = sorted(vendors, key=lambda x: x[1])

        prices = [v[1] for v in vendors_sorted]
        unit = vendors_sorted[0][2] if vendors_sorted[0][2] else ""
        unit_str = f"/{unit}" if unit else ""

        low = min(prices)
        high = max(prices)
        cheapest_vendor = vendors_sorted[0][0]

        lines.append(f"  {item.upper()}")
        if low == high:
            lines.append(f"  Price: ${low:.2f}{unit_str} (all vendors same price)")
        else:
            lines.append(f"  Range: ${low:.2f} - ${high:.2f}{unit_str}  |  Cheapest: {cheapest_vendor}")
        lines.append("")

        for vendor, price, unit in vendors_sorted:
            u = f"/{unit}" if unit else ""
            lines.append(f"    ${price:.2f}{u}  —  {vendor}")

        lines.append("")
        lines.append("-" * 50)
        lines.append("")

    lines.append("— Farmlind Availability Bot")
    return "\n".join(lines)


# ── Spreadsheet builder ────────────────────────────────────────────────────────

GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)


def normalize_unit(unit):
    """Normalize unit strings so near-duplicates collapse (e.g. '6 ct' → '6ct')."""
    if not unit:
        return ""
    u = unit.strip().lower()
    u = re.sub(r"\s+", "", u)          # remove all spaces: "6 ct" → "6ct"
    u = re.sub(r"s$", "", u)           # strip trailing s: "lbs" → "lb", "bags" → "bag"
    return u


def build_spreadsheet(item_vendor_map):
    """Matrix layout: vendors across top, items down side.
    Each cell shows deduplicated price/unit combos for that vendor, one per line,
    sorted cheapest first — e.g.:
        $2.00 / 6ct
        $3.50 / 12ct
    Green = cheapest vendor, Red = most expensive, Yellow = middle."""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Price Comparison"

    # Alternating row fills for readability
    ROW_FILL_ODD  = PatternFill("solid", fgColor="FFFFFF")
    ROW_FILL_EVEN = PatternFill("solid", fgColor="F2F2F2")

    # Collect vendors, sort alphabetically with Specials at end
    all_vendors = set()
    for entries in item_vendor_map.values():
        for vendor, _price, _unit in entries:
            all_vendors.add(vendor)
    def vendor_sort_key(v):
        return (1 if "specials" in v.lower() else 0, v.lower())
    vendors = sorted(all_vendors, key=vendor_sort_key)

    # Header row: Item | Cheapest | Price Range | Vendor1 | Vendor2 | ...
    ws.cell(row=1, column=1, value="Item")
    ws.cell(row=1, column=2, value="Cheapest Vendor")
    ws.cell(row=1, column=3, value="Price Range")
    vendor_col = {}
    for idx, vendor in enumerate(vendors):
        col = 4 + idx
        vendor_col[vendor] = col
        ws.cell(row=1, column=col, value=vendor)

    last_col = 3 + len(vendors)
    for c in range(1, last_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    for vendor in vendors:
        ws.column_dimensions[openpyxl.utils.get_column_letter(vendor_col[vendor])].width = 16

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "D2"

    # Item rows
    row_idx = 2
    for item in sorted(item_vendor_map.keys()):
        entries = item_vendor_map[item]

        # Group by vendor, deduplicating by (price, normalized_unit)
        per_vendor = {}
        for vendor, price, unit in entries:
            norm_u = normalize_unit(unit)
            if vendor not in per_vendor:
                per_vendor[vendor] = {}
            key = (round(price, 4), norm_u)
            # Keep the original unit string from the first time we see this combo
            if key not in per_vendor[vendor]:
                per_vendor[vendor][key] = (price, unit.strip() if unit else "")

        # Find global min/max prices for coloring
        all_prices = [price for slots in per_vendor.values() for price, _ in slots.values()]
        low, high = min(all_prices), max(all_prices)

        # Cheapest vendor = one whose lowest price == global low
        cheapest_vendor = min(
            per_vendor.items(),
            key=lambda kv: min(p for p, _ in kv[1].values())
        )[0]
        price_range = f"${low:.2f} — ${high:.2f}" if low != high else f"${low:.2f}"

        # Alternating row background
        row_bg = ROW_FILL_ODD if (row_idx % 2 == 0) else ROW_FILL_EVEN

        # Item name cell
        item_cell = ws.cell(row=row_idx, column=1, value=item)
        item_cell.font = Font(bold=True, size=11)
        item_cell.alignment = Alignment(horizontal="left", vertical="center")
        item_cell.border = THIN_BORDER
        item_cell.fill = row_bg

        # Cheapest vendor cell
        cheap_cell = ws.cell(row=row_idx, column=2, value=cheapest_vendor)
        cheap_cell.alignment = Alignment(horizontal="center", vertical="center")
        cheap_cell.border = THIN_BORDER
        cheap_cell.font = Font(bold=True, color="1F6B23")
        cheap_cell.fill = row_bg

        # Price range cell
        range_cell = ws.cell(row=row_idx, column=3, value=price_range)
        range_cell.alignment = Alignment(horizontal="center", vertical="center")
        range_cell.border = THIN_BORDER
        range_cell.fill = row_bg

        # Vendor columns
        for vendor in vendors:
            col = vendor_col[vendor]
            cell = ws.cell(row=row_idx, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            slot = per_vendor.get(vendor)
            if not slot:
                cell.fill = row_bg
                continue

            # Sort deduplicated variants cheapest first, build display lines
            sorted_variants = sorted(slot.values(), key=lambda x: x[0])
            display_parts = []
            for price, unit in sorted_variants:
                if unit:
                    display_parts.append(f"${price:.2f} / {unit}")
                else:
                    display_parts.append(f"${price:.2f}")
            cell.value = "\n".join(display_parts)
            cell.font = Font(size=11)

            # Color by whether this vendor's cheapest price is the global low/high
            vendor_min = min(p for p, _ in slot.values())
            if vendor_min == low:
                cell.fill = GREEN_FILL
            elif vendor_min == high and low != high:
                cell.fill = RED_FILL
            else:
                cell.fill = YELLOW_FILL

        # Auto-size row height based on max lines in any vendor cell for this row
        max_lines = max(
            len(list(per_vendor.get(v, {}).values())) for v in vendors
        ) if vendors else 1
        ws.row_dimensions[row_idx].height = max(18, max_lines * 18)

        row_idx += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Item name normalization ────────────────────────────────────────────────────

import re
import string

# Specials can be offered any day of the week. Detection in the spreadsheet uses the
# substring "specials" (case-insensitive), so dynamic labels like "Sunday Specials - TMK"
# are still recognized; this set is kept for reference.
SPECIALS_VENDORS = {f"{day} Specials" for day in
                    ("Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday")}

def load_item_aliases():
  """Load item aliases from external JSON file."""
  aliases_file = os.path.join(os.path.dirname(__file__), "item_aliases.json")
  if not os.path.exists(aliases_file):
    print("Warning: item_aliases.json not found — using empty aliases")
    return {}
  try:
    with open(aliases_file, 'r') as f:
      data = json.load(f)
    flat = {}
    for category in data.values():
      if isinstance(category, dict):
        flat.update(category)
    return flat
  except Exception as e:
    print(f"Warning: could not load item_aliases.json: {e}")
    return {}

ITEM_ALIASES = load_item_aliases()

def normalize_item_name(name):
    """Normalize item names so minor variations map to the same key."""
    name = re.sub(r"[,\.\/\\]", " ", name)
    name = " ".join(name.split())
    lower = name.lower()
    if lower in ITEM_ALIASES:
        return ITEM_ALIASES[lower]
    return name.title()


STATIC_LISTS_DIR = os.path.join(os.path.dirname(__file__), "static_price_lists")


def merge_result_into_map(result, item_vendor_map, item_display_name, seen_vendors,
                          vendor_item_keys, fill_only=False):
    """Merge one extraction result into the running item->vendor map.

    fill_only=False (Slack): the freshest source for a vendor. Skipped entirely if that
      vendor was already merged (older message for the same vendor).
    fill_only=True (static baseline): only adds items for keys this vendor hasn't already
      provided from a fresher source. So when Nathel posts in Slack, only the items in that
      post are replaced; items not re-posted keep their baseline price.

    Returns the vendor name if processed, or None if skipped."""
    # If the extraction split into multiple vendor groups (e.g. a Nathel screenshot with
    # both Josh's and Paul's orders), merge each group separately and ignore the top-level.
    groups = result.get("groups")
    if groups:
        last = None
        for g in groups:
            if g and g.get("vendor") and g.get("items"):
                last = merge_result_into_map(g, item_vendor_map, item_display_name,
                                             seen_vendors, vendor_item_keys, fill_only=fill_only)
        return last

    vendor = result.get("vendor")
    if not vendor:
        print("    Skipping — no vendor name found in content")
        return None

    # Farmlind / Matt are the BUYER, never a vendor — drop them outright.
    # Hunts Point is a market, not a vendor.
    low_v = vendor.lower()
    if any(bad in low_v for bad in ("farmlind", "matt lind", "matt ", "hunts point")) or low_v.strip() == "matt":
        print(f"    Skipping '{vendor}' — not a vendor")
        return None

    if not fill_only:
        # Slack: only the most recent message per vendor counts
        if vendor in seen_vendors:
            print(f"    Skipping older {vendor} source")
            return None
        seen_vendors.add(vendor)

    # Snapshot the keys this vendor already had BEFORE this merge. In fill-only
    # (baseline) mode we skip only those — i.e. items a fresher Slack post already
    # supplied — while still allowing this list's own size variants of the same item.
    prior_keys = set(vendor_item_keys[vendor])

    items = result.get("items", [])
    added = 0
    skipped = 0
    for entry in items:
        item = entry.get("item", "").strip()
        price = entry.get("price")
        unit = entry.get("unit", "") or ""
        if not item or price is None:
            continue
        try:
            key = normalize_item_name(item)
        except (ValueError, TypeError):
            continue
        # In fill-only mode, don't overwrite an item this vendor already supplied
        # from a fresher (Slack) source.
        if fill_only and key in prior_keys:
            skipped += 1
            continue
        try:
            price_f = float(price)
        except (ValueError, TypeError):
            continue
        if key not in item_display_name:
            item_display_name[key] = item
        row = (vendor, price_f, unit.strip())
        if row not in item_vendor_map[key]:
            item_vendor_map[key].append(row)
        vendor_item_keys[vendor].add(key)
        added += 1

    if fill_only:
        print(f"    Vendor: {vendor} — {added} baseline items added, {skipped} kept fresh from Slack")
    else:
        print(f"    Vendor: {vendor} — {added} items")
    return vendor


def process_static_lists(item_vendor_map, item_display_name, seen_vendors,
                         vendor_item_keys, product_names):
    """Parse any baseline price lists bundled in static_price_lists/ and merge them in.
    These are vendors whose lists are NOT posted to Slack. A newer Slack post for the
    same vendor (processed earlier) takes priority via seen_vendors."""
    if not os.path.isdir(STATIC_LISTS_DIR):
        return
    for filename in sorted(os.listdir(STATIC_LISTS_DIR)):
        path = os.path.join(STATIC_LISTS_DIR, filename)
        if not os.path.isfile(path):
            continue
        lower = filename.lower()
        try:
            with open(path, "rb") as fh:
                file_bytes = fh.read()

            print(f"  Parsing static price list: {filename}...")
            if lower.endswith(".pdf"):
                # Multi-page price lists are processed page-by-page so nothing is
                # truncated by the model output limit.
                result = extract_prices_from_pdf_chunked(file_bytes, product_names)
            elif lower.endswith((".png", ".jpg", ".jpeg", ".gif", ".webp")):
                ext = lower.rsplit(".", 1)[-1]
                media = "image/jpeg" if ext in ("jpg", "jpeg") else f"image/{ext}"
                block = {
                    "type": "image",
                    "source": {"type": "base64", "media_type": media,
                               "data": base64.b64encode(file_bytes).decode()}
                }
                result = extract_prices_from_content([block], "Static List", product_names)
            elif lower.endswith((".xlsx", ".xls")):
                import openpyxl as _oxl
                wb = _oxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
                text_content = "\n".join(
                    " ".join(str(c) for c in row if c is not None)
                    for sheet in wb.worksheets
                    for row in sheet.iter_rows(values_only=True)
                )
                result = extract_prices_from_content(
                    [{"type": "text", "text": text_content}], "Static List", product_names
                )
            else:
                continue

            if result:
                merge_result_into_map(result, item_vendor_map, item_display_name,
                                      seen_vendors, vendor_item_keys, fill_only=True)
        except Exception as e:
            print(f"  Static list error ({filename}): {e}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print("Fetching availability price data from Slack...")

    print("Loading product reference...")
    product_names = load_product_reference()

    channel_id = find_channel_id(CHANNEL_NAME)
    print(f"Found channel #{CHANNEL_NAME}: {channel_id}")

    messages = fetch_channel_history(channel_id, limit=200)
    print(f"Fetched {len(messages)} messages")

    from datetime import timezone, timedelta
    et = timezone(timedelta(hours=-4))
    now_et = datetime.now(et)
    today_et = now_et.date()
    today_weekday = now_et.weekday()  # 1=Tuesday, 5=Saturday
    is_scheduled_day = today_weekday in (1, 5)
    run_mode = os.environ.get("RUN_MODE", "daily")
    force_send = bool(os.environ.get("FORCE_SEND"))

    # Fetch Dottavio's most recent price-list email (sent ~Thursday mornings).
    print("Checking for Dottavio price-list email...")
    dottavio_text, dottavio_ts = fetch_latest_dottavio_email()
    dottavio_today = False
    if dottavio_ts:
        d_date = datetime.fromtimestamp(dottavio_ts / 1000, tz=et).date()
        dottavio_today = (d_date == today_et)
        print(f"  Dottavio email found, dated {d_date} (today={dottavio_today})")
    else:
        print("  No recent Dottavio email found")

    # ── Decide whether to send ──
    if force_send:
        should_send = True
    elif run_mode == "dottavio_watch":
        # Thursday-morning watch: only send when a NEW Dottavio list arrived today
        # and we haven't already sent today's spreadsheet.
        if dottavio_today and not availability_already_sent_today(today_et):
            should_send = True
        else:
            should_send = False
            print("Dottavio watch: nothing new to send (already sent today or no new list).")
    else:  # daily 8pm run
        should_send = is_scheduled_day or has_messages_today(messages) or dottavio_today

    if not should_send:
        print("Not a scheduled send and nothing new today — skipping email.")
        return

    # item -> list of (vendor, price, unit)
    # Keys are normalized item names; values track (vendor, price, unit)
    item_vendor_map = defaultdict(list)
    # canonical name lookup: normalized_key -> display name
    item_display_name = {}
    # track which vendors we've already processed (specials only use most recent)
    seen_vendors = set()
    # vendor -> set of normalized item keys already supplied (Slack takes priority,
    # static baseline only fills the gaps)
    vendor_item_keys = defaultdict(set)

    for msg in messages:
        ts = msg.get("ts", "0")
        user_id = msg.get("user", "")
        sender_name = get_user_display_name(user_id) if user_id else "Unknown"

        # Non-PDF content (text + images + spreadsheet text) goes in one bundle;
        # PDFs are extracted page-by-page so big multi-page lists aren't truncated.
        content_blocks = []
        pdf_files = []

        text = msg.get("text", "").strip()
        if text:
            content_blocks.append({"type": "text", "text": text})

        for f in msg.get("files", []):
            mime = f.get("mimetype", "")
            url = f.get("url_private_download") or f.get("url_private")
            filename = f.get("name", "")
            if not url:
                continue
            print(f"  Downloading: {filename} ({mime})")
            try:
                file_bytes = download_slack_file(url)
                if "pdf" in mime or filename.lower().endswith(".pdf"):
                    pdf_files.append((filename, file_bytes))
                elif mime.startswith("image/"):
                    content_blocks.append({
                        "type": "image",
                        "source": {"type": "base64", "media_type": mime,
                                   "data": base64.b64encode(file_bytes).decode()}
                    })
                elif filename.lower().endswith((".xlsx", ".xls")):
                    import openpyxl as _oxl
                    wb = _oxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
                    text_content = "\n".join(
                        " ".join(str(c) for c in row if c is not None)
                        for sheet in wb.worksheets
                        for row in sheet.iter_rows(values_only=True)
                    )
                    content_blocks.append({"type": "text", "text": text_content})
            except Exception as e:
                print(f"  File error: {e}")

        # A message with only a bare "uploaded a file" stub and no real content is skipped
        has_real_text = any(b.get("type") == "text" and b.get("text", "").strip()
                            for b in content_blocks)
        if not pdf_files and not has_real_text and not any(
                b.get("type") == "image" for b in content_blocks):
            continue

        msg_date = datetime.fromtimestamp(float(ts)).strftime("%Y-%m-%d")
        print(f"  Parsing message from {sender_name} ({msg_date})...")
        try:
            # PDFs first (the actual price lists), then the text/image bundle.
            for fname, fbytes in pdf_files:
                print(f"  PDF: {fname}")
                pres = extract_prices_from_pdf_chunked(
                    fbytes, product_names,
                    seen_check=lambda v: v in seen_vendors
                )
                if pres:
                    merge_result_into_map(pres, item_vendor_map, item_display_name,
                                          seen_vendors, vendor_item_keys)
            if has_real_text or any(b.get("type") == "image" for b in content_blocks):
                result = extract_prices_from_content(content_blocks, sender_name, product_names)
                if result:
                    merge_result_into_map(result, item_vendor_map, item_display_name,
                                          seen_vendors, vendor_item_keys)
        except Exception as e:
            print(f"  Parse error: {e}")

    # Merge in Dottavio's emailed price list (vendor name isn't in the body, so force it).
    if dottavio_text:
        print("\nParsing Dottavio price-list email...")
        try:
            d_result = extract_prices_from_content(
                [{"type": "text", "text": dottavio_text}], "Dottavio", product_names
            )
            if d_result:
                d_result["vendor"] = "Dottavio"
                merge_result_into_map(d_result, item_vendor_map, item_display_name,
                                      seen_vendors, vendor_item_keys)
        except Exception as e:
            print(f"  Dottavio parse error: {e}")

    # Merge in baseline price lists not posted to Slack (e.g. Nathel).
    # Slack was processed first, so a newer Slack post for the same vendor wins.
    print("\nProcessing static (non-Slack) price lists...")
    process_static_lists(item_vendor_map, item_display_name, seen_vendors,
                         vendor_item_keys, product_names)

    if not item_vendor_map:
        print("No price data found — nothing to send.")
        return

    # Rebuild map using display names for output
    display_map = defaultdict(list)
    for key, entries in item_vendor_map.items():
        display_name = item_display_name.get(key, key)
        display_map[display_name] = entries

    print(f"\nTotal unique items: {len(display_map)}")

    today_str = datetime.now().strftime("%B %d, %Y")
    subject = f"Availability Price Comparison — {today_str}"

    print("Building spreadsheet...")
    xlsx_bytes = build_spreadsheet(display_map)

    body = (
        f"Hi Matt,\n\n"
        f"Attached is today's availability price comparison across all vendors.\n\n"
        f"GREEN = cheapest option   |   RED = most expensive   |   YELLOW = middle\n\n"
        f"— Farmlind Availability Bot"
    )

    print("Sending email...")
    token = get_access_token()
    send_email_with_attachment(token, subject, body, xlsx_bytes)
    print("Done — email sent with spreadsheet attached.")


if __name__ == "__main__":
    main()
